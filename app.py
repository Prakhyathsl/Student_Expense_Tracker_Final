from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
import os
from datetime import datetime
from threading import RLock

from database import init_db as init_database, get_expenses as db_get_expenses, save_expenses as db_save_expenses, read_savings_data as db_read_savings_data, save_savings_data as db_save_savings_data


# ============================================================
# APPLICATION
# ============================================================

app = Flask(__name__)


# ============================================================
# PATHS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent

DATA_DIR = BASE_DIR / "data"

DATA_DIR.mkdir(parents=True, exist_ok=True)

# New master file used by the application
MASTER_EXCEL = DATA_DIR / "Expense_Tracker_Master.xlsx"

# Savings are stored separately so the existing
# expense system is not disturbed.
SAVINGS_EXCEL = DATA_DIR / "Savings_Tracker.xlsx"

SAVINGS_GOALS_SHEET = "Goals"
SAVINGS_HISTORY_SHEET = "Savings"

# Your old records
OLD_EXCEL = DATA_DIR / "Student_Expense_Tracker.xlsx"

SHEET_NAME = "Expenses"


# ============================================================
# FILE LOCK
# ============================================================

# RLock is used because some functions call other functions
# which also need the lock.
FILE_LOCK = RLock()


# ============================================================
# EXCEL HEADERS
# ============================================================

HEADERS = [
    "ID",
    "Academic Year",
    "Semester",
    "Date",
    "Category",
    "Amount",
    "Payment Method",
    "Description"
]


# ============================================================
# ACADEMIC YEARS
# ============================================================

ACADEMIC_YEARS = [
    "1st Year",
    "2nd Year",
    "3rd Year",
    "4th Year"
]


# ============================================================
# SEMESTERS
# ============================================================

SEMESTERS = [
    "1st Sem",
    "2nd Sem",
    "3rd Sem",
    "4th Sem",
    "5th Sem",
    "6th Sem",
    "7th Sem",
    "8th Sem"
]


# ============================================================
# YEAR → SEMESTER MAPPING
# ============================================================

YEAR_SEMESTERS = {
    "1st Year": [
        "1st Sem",
        "2nd Sem"
    ],

    "2nd Year": [
        "3rd Sem",
        "4th Sem"
    ],

    "3rd Year": [
        "5th Sem",
        "6th Sem"
    ],

    "4th Year": [
        "7th Sem",
        "8th Sem"
    ]
}


# ============================================================
# DEFAULT CATEGORIES
# ============================================================

CATEGORIES = [
    "Food",
    "Transport",
    "Education",
    "Shopping",
    "Entertainment",
    "Bills",
    "Health",
    "Travel",
    "Other"
]


# ============================================================
# PAYMENT METHODS
# ============================================================

PAYMENT_METHODS = [
    "Cash",
    "UPI",
    "Card",
    "Bank Transfer"
]


# ============================================================
# DATE HELPERS
# ============================================================

def normalize_date(value):
    """
    Convert supported date formats into YYYY-MM-DD.

    Supported:
        DD-MM-YYYY
        DD/MM/YYYY
        YYYY-MM-DD
        YYYY/MM/DD
        Python datetime
    """

    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    text = str(value).strip()

    if not text:
        return ""

    formats = [
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%Y/%m/%d"
    ]

    for date_format in formats:
        try:
            date = datetime.strptime(
                text,
                date_format
            )

            return date.strftime(
                "%Y-%m-%d"
            )

        except ValueError:
            continue

    return ""


def display_date(value):
    """
    Convert stored YYYY-MM-DD into DD-MM-YYYY.
    """

    normalized = normalize_date(value)

    if not normalized:
        return ""

    try:
        date = datetime.strptime(
            normalized,
            "%Y-%m-%d"
        )

        return date.strftime(
            "%d-%m-%Y"
        )

    except ValueError:
        return ""


# ============================================================
# EXCEL FORMATTING
# ============================================================

def format_excel(worksheet):
    """
    Apply formatting to the Excel worksheet.
    """

    # Header
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    for cell in worksheet[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )


    # Freeze header
    worksheet.freeze_panes = "A2"


    # Column widths
    widths = {
        "A": 8,
        "B": 18,
        "C": 14,
        "D": 15,
        "E": 25,
        "F": 16,
        "G": 20,
        "H": 45
    }


    for column, width in widths.items():

        worksheet.column_dimensions[
            column
        ].width = width


    # Date and amount formatting
    for row in range(
        2,
        worksheet.max_row + 1
    ):

        date_cell = worksheet.cell(
            row=row,
            column=4
        )

        amount_cell = worksheet.cell(
            row=row,
            column=6
        )


        if isinstance(
            date_cell.value,
            datetime
        ):

            date_cell.number_format = (
                "dd-mm-yyyy"
            )


        amount_cell.number_format = (
            '₹#,##0.00'
        )


# ============================================================
# CREATE MASTER EXCEL
# ============================================================

def create_master_file():
    """
    Create a new empty master Excel file.
    """

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = SHEET_NAME

    worksheet.append(
        HEADERS
    )

    format_excel(
        worksheet
    )

    workbook.save(
        MASTER_EXCEL
    )


# ============================================================
# IMPORT OLD EXCEL
# ============================================================

def import_old_records():
    """
    If the master file doesn't exist, import the old
    Student_Expense_Tracker.xlsx file.

    This preserves the user's old records.
    """

    if MASTER_EXCEL.exists():
        return


    if not OLD_EXCEL.exists():

        create_master_file()

        return


    try:

        old_workbook = load_workbook(
            OLD_EXCEL,
            data_only=True
        )

        old_worksheet = (
            old_workbook.active
        )


        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = SHEET_NAME

        worksheet.append(
            HEADERS
        )


        next_id = 1


        for row in old_worksheet.iter_rows(
            min_row=2,
            values_only=True
        ):

            # Ignore completely empty rows
            if not any(
                value is not None
                for value in row
            ):
                continue


            values = list(row)


            while len(values) < 8:
                values.append(None)


            # ------------------------------------------------
            # ID
            # ------------------------------------------------

            try:

                record_id = int(
                    values[0]
                )

            except (
                TypeError,
                ValueError
            ):

                record_id = next_id


            next_id = max(
                next_id,
                record_id + 1
            )


            # ------------------------------------------------
            # ACADEMIC YEAR
            # ------------------------------------------------

            academic_year = str(
                values[1] or ""
            ).strip()


            # ------------------------------------------------
            # SEMESTER
            # ------------------------------------------------

            semester = str(
                values[2] or ""
            ).strip()


            # ------------------------------------------------
            # DATE
            # ------------------------------------------------

            date_text = normalize_date(
                values[3]
            )


            excel_date = ""

            if date_text:

                try:

                    excel_date = datetime.strptime(
                        date_text,
                        "%Y-%m-%d"
                    )

                except ValueError:

                    excel_date = ""


            # ------------------------------------------------
            # CATEGORY
            # ------------------------------------------------

            category = str(
                values[4] or ""
            ).strip()


            # ------------------------------------------------
            # AMOUNT
            # ------------------------------------------------

            try:

                amount = float(
                    values[5] or 0
                )

            except (
                TypeError,
                ValueError
            ):

                amount = 0.0


            # ------------------------------------------------
            # PAYMENT METHOD
            # ------------------------------------------------

            payment_method = str(
                values[6] or ""
            ).strip()


            # ------------------------------------------------
            # DESCRIPTION
            # ------------------------------------------------

            description = str(
                values[7] or ""
            ).strip()


            worksheet.append([
                record_id,
                academic_year,
                semester,
                excel_date,
                category,
                amount,
                payment_method,
                description
            ])


        format_excel(
            worksheet
        )


        workbook.save(
            MASTER_EXCEL
        )


        print(
            f"Imported old records from: {OLD_EXCEL}"
        )


    except Exception as error:

        print(
            "ERROR IMPORTING OLD EXCEL:",
            error
        )


        # If importing fails, create empty master
        create_master_file()


# ============================================================
# INITIALIZE DATA
# ============================================================

def initialize_data():

    with FILE_LOCK:

        if MASTER_EXCEL.exists():
            return

        import_old_records()


# ============================================================
# READ ALL EXPENSES
# ============================================================

def get_expenses():

    initialize_data()


    with FILE_LOCK:

        workbook = load_workbook(
            MASTER_EXCEL,
            data_only=True
        )

        if SHEET_NAME not in workbook.sheetnames:

            workbook.close()

            create_master_file()

            return []


        worksheet = workbook[
            SHEET_NAME
        ]


        expenses = []


        for row in worksheet.iter_rows(
            min_row=2,
            values_only=True
        ):

            if not any(
                value is not None
                for value in row
            ):
                continue


            values = list(row)


            while len(values) < 8:
                values.append(None)


            try:

                expense_id = int(
                    values[0]
                )

            except (
                TypeError,
                ValueError
            ):

                continue


            try:

                amount = float(
                    values[5] or 0
                )

            except (
                TypeError,
                ValueError
            ):

                amount = 0.0


            expense = {

                "id":
                    expense_id,

                "academic_year":
                    str(
                        values[1] or ""
                    ).strip(),

                "semester":
                    str(
                        values[2] or ""
                    ).strip(),

                "date":
                    normalize_date(
                        values[3]
                    ),

                "category":
                    str(
                        values[4] or ""
                    ).strip(),

                "amount":
                    amount,

                "payment_method":
                    str(
                        values[6] or ""
                    ).strip(),

                "description":
                    str(
                        values[7] or ""
                    ).strip()

            }


            expenses.append(
                expense
            )


        workbook.close()


        return expenses


# ============================================================
# SAVE ALL EXPENSES
# ============================================================

def save_expenses(expenses):

    with FILE_LOCK:

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = SHEET_NAME

        worksheet.append(
            HEADERS
        )


        for expense in expenses:

            date_text = normalize_date(
                expense.get(
                    "date",
                    ""
                )
            )


            excel_date = ""

            if date_text:

                try:

                    excel_date = datetime.strptime(
                        date_text,
                        "%Y-%m-%d"
                    )

                except ValueError:

                    excel_date = ""


            worksheet.append([

                int(
                    expense.get(
                        "id",
                        0
                    )
                ),

                str(
                    expense.get(
                        "academic_year",
                        ""
                    )
                ),

                str(
                    expense.get(
                        "semester",
                        ""
                    )
                ),

                excel_date,

                str(
                    expense.get(
                        "category",
                        ""
                    )
                ),

                float(
                    expense.get(
                        "amount",
                        0
                    )
                ),

                str(
                    expense.get(
                        "payment_method",
                        ""
                    )
                ),

                str(
                    expense.get(
                        "description",
                        ""
                    )
                )

            ])


        format_excel(
            worksheet
        )


        temporary_file = (
            DATA_DIR /
            "Expense_Tracker_Master_temp.xlsx"
        )


        workbook.save(
            temporary_file
        )

        workbook.close()


        # Replace master safely
        temporary_file.replace(
            MASTER_EXCEL
        )


# ============================================================
# HOME PAGE
# ============================================================

@app.route("/")
def home():

    return render_template(
        "index.html",
        categories=CATEGORIES,
        payment_methods=PAYMENT_METHODS,
        academic_years=ACADEMIC_YEARS,
        semesters=SEMESTERS,
        year_semesters=YEAR_SEMESTERS
    )


# ============================================================
# GET EXPENSES
# ============================================================

@app.route(
    "/api/expenses",
    methods=["GET"]
)
def api_get_expenses():

    try:

        expenses = get_expenses()


        return jsonify(
            expenses
        )


    except Exception as error:

        print(
            "GET EXPENSES ERROR:",
            error
        )


        return jsonify({
            "error": str(error)
        }), 500


# ============================================================
# ADD EXPENSE
# ============================================================

@app.route(
    "/api/expenses",
    methods=["POST"]
)
def api_add_expense():

    try:

        data = request.get_json(
            silent=True
        )


        if not isinstance(
            data,
            dict
        ):

            return jsonify({
                "error":
                    "Invalid request data."
            }), 400


        # ----------------------------------------------------
        # DATA
        # ----------------------------------------------------

        academic_year = str(
            data.get(
                "academic_year",
                ""
            )
        ).strip()


        semester = str(
            data.get(
                "semester",
                ""
            )
        ).strip()


        date = normalize_date(
            data.get(
                "date",
                ""
            )
        )


        category = str(
            data.get(
                "category",
                ""
            )
        ).strip()


        payment_method = str(
            data.get(
                "payment_method",
                ""
            )
        ).strip()


        description = str(
            data.get(
                "description",
                ""
            )
        ).strip()


        # ----------------------------------------------------
        # VALIDATE YEAR
        # ----------------------------------------------------

        if academic_year not in ACADEMIC_YEARS:

            return jsonify({
                "error":
                    "Please select a valid academic year."
            }), 400


        # ----------------------------------------------------
        # VALIDATE SEMESTER
        # ----------------------------------------------------

        valid_semesters = (
            YEAR_SEMESTERS.get(
                academic_year,
                []
            )
        )


        if semester not in valid_semesters:

            return jsonify({
                "error":
                    "Please select a valid semester for the selected academic year."
            }), 400


        # ----------------------------------------------------
        # VALIDATE DATE
        # ----------------------------------------------------

        if not date:

            return jsonify({
                "error":
                    "Please enter a valid date in DD-MM-YYYY format."
            }), 400


        try:

            datetime.strptime(
                date,
                "%Y-%m-%d"
            )

        except ValueError:

            return jsonify({
                "error":
                    "Invalid date."
            }), 400


        # ----------------------------------------------------
        # VALIDATE CATEGORY
        # ----------------------------------------------------

        if not category:

            return jsonify({
                "error":
                    "Please enter a category."
            }), 400


        # ----------------------------------------------------
        # AMOUNT
        # ----------------------------------------------------

        try:

            amount = float(
                data.get(
                    "amount",
                    0
                )
            )

        except (
            TypeError,
            ValueError
        ):

            return jsonify({
                "error":
                    "Please enter a valid amount."
            }), 400


        if amount < 0:

            return jsonify({
                "error":
                    "Amount cannot be negative."
            }), 400


        # ----------------------------------------------------
        # PAYMENT METHOD
        # ----------------------------------------------------

        if payment_method not in PAYMENT_METHODS:

            return jsonify({
                "error":
                    "Please select a valid payment method."
            }), 400


        # ----------------------------------------------------
        # SAVE
        # ----------------------------------------------------

        expenses = get_expenses()


        if expenses:

            new_id = max(
                expense["id"]
                for expense in expenses
            ) + 1

        else:

            new_id = 1


        new_expense = {

            "id":
                new_id,

            "academic_year":
                academic_year,

            "semester":
                semester,

            "date":
                date,

            "category":
                category,

            "amount":
                amount,

            "payment_method":
                payment_method,

            "description":
                description

        }


        expenses.append(
            new_expense
        )


        save_expenses(
            expenses
        )


        return jsonify({

            "success":
                True,

            "expense":
                new_expense

        }), 201


    except Exception as error:

        print(
            "ADD EXPENSE ERROR:",
            error
        )


        return jsonify({
            "error": str(error)
        }), 500


# ============================================================
# UPDATE EXPENSE
# ============================================================

@app.route(
    "/api/expenses/<int:expense_id>",
    methods=["PUT"]
)
def api_update_expense(
    expense_id
):

    try:

        data = request.get_json(
            silent=True
        ) or {}


        expenses = get_expenses()


        expense = None


        for item in expenses:

            if item["id"] == expense_id:

                expense = item

                break


        if expense is None:

            return jsonify({
                "error":
                    "Expense not found."
            }), 404


        # ----------------------------------------------------
        # UPDATED VALUES
        # ----------------------------------------------------

        academic_year = str(
            data.get(
                "academic_year",
                expense["academic_year"]
            )
        ).strip()


        semester = str(
            data.get(
                "semester",
                expense["semester"]
            )
        ).strip()


        date = normalize_date(
            data.get(
                "date",
                expense["date"]
            )
        )


        category = str(
            data.get(
                "category",
                expense["category"]
            )
        ).strip()


        payment_method = str(
            data.get(
                "payment_method",
                expense["payment_method"]
            )
        ).strip()


        description = str(
            data.get(
                "description",
                expense["description"]
            )
        ).strip()


        try:

            amount = float(
                data.get(
                    "amount",
                    expense["amount"]
                )
            )

        except (
            TypeError,
            ValueError
        ):

            return jsonify({
                "error":
                    "Invalid amount."
            }), 400


        # ----------------------------------------------------
        # VALIDATION
        # ----------------------------------------------------

        if academic_year not in ACADEMIC_YEARS:

            return jsonify({
                "error":
                    "Invalid academic year."
            }), 400


        if semester not in YEAR_SEMESTERS.get(
            academic_year,
            []
        ):

            return jsonify({
                "error":
                    "Invalid semester."
            }), 400


        if not date:

            return jsonify({
                "error":
                    "Invalid date."
            }), 400


        if not category:

            return jsonify({
                "error":
                    "Category is required."
            }), 400


        if amount < 0:

            return jsonify({
                "error":
                    "Amount cannot be negative."
            }), 400


        if payment_method not in PAYMENT_METHODS:

            return jsonify({
                "error":
                    "Invalid payment method."
            }), 400


        # ----------------------------------------------------
        # UPDATE
        # ----------------------------------------------------

        expense["academic_year"] = (
            academic_year
        )

        expense["semester"] = (
            semester
        )

        expense["date"] = (
            date
        )

        expense["category"] = (
            category
        )

        expense["amount"] = (
            amount
        )

        expense["payment_method"] = (
            payment_method
        )

        expense["description"] = (
            description
        )


        save_expenses(
            expenses
        )


        return jsonify({

            "success":
                True,

            "expense":
                expense

        })


    except Exception as error:

        print(
            "UPDATE EXPENSE ERROR:",
            error
        )


        return jsonify({
            "error": str(error)
        }), 500


# ============================================================
# DELETE EXPENSE
# ============================================================

@app.route(
    "/api/expenses/<int:expense_id>",
    methods=["DELETE"]
)
def api_delete_expense(
    expense_id
):

    try:

        expenses = get_expenses()


        original_count = len(
            expenses
        )


        remaining_expenses = [

            expense

            for expense in expenses

            if expense["id"] != expense_id

        ]


        if len(
            remaining_expenses
        ) == original_count:

            return jsonify({
                "error":
                    "Expense not found."
            }), 404


        save_expenses(
            remaining_expenses
        )


        return jsonify({
            "success": True
        })


    except Exception as error:

        print(
            "DELETE EXPENSE ERROR:",
            error
        )


        return jsonify({
            "error": str(error)
        }), 500


# ============================================================
# DASHBOARD STATISTICS
# ============================================================

@app.route(
    "/api/stats",
    methods=["GET"]
)
def api_stats():

    try:

        expenses = get_expenses()


        # ----------------------------------------------------
        # TOTAL
        # ----------------------------------------------------

        total_expenses = sum(
            expense["amount"]
            for expense in expenses
        )


        transaction_count = len(
            expenses
        )


        average_expense = (

            total_expenses /
            transaction_count

            if transaction_count > 0

            else 0

        )


        # ----------------------------------------------------
        # CATEGORY TOTALS
        # ----------------------------------------------------

        category_totals = {}


        for expense in expenses:

            category = (
                expense["category"]
                or "Other"
            )


            category_totals[
                category
            ] = (

                category_totals.get(
                    category,
                    0
                )

                +

                expense["amount"]

            )


        # ----------------------------------------------------
        # ACADEMIC YEAR TOTALS
        # ----------------------------------------------------

        year_totals = {

            year: 0

            for year in
            ACADEMIC_YEARS

        }


        for expense in expenses:

            year = (
                expense["academic_year"]
            )


            if year in year_totals:

                year_totals[
                    year
                ] += expense["amount"]


        # ----------------------------------------------------
        # SEMESTER TOTALS
        # ----------------------------------------------------

        semester_totals = {

            semester: 0

            for semester in
            SEMESTERS

        }


        for expense in expenses:

            semester = (
                expense["semester"]
            )


            if semester in semester_totals:

                semester_totals[
                    semester
                ] += expense["amount"]


        # ----------------------------------------------------
        # PAYMENT TOTALS
        # ----------------------------------------------------

        payment_totals = {}


        for expense in expenses:

            payment = (
                expense["payment_method"]
                or "Unknown"
            )


            payment_totals[
                payment
            ] = (

                payment_totals.get(
                    payment,
                    0
                )

                +

                expense["amount"]

            )


        # ----------------------------------------------------
        # THIS MONTH
        # ----------------------------------------------------

        now = datetime.now()

        this_month = 0


        for expense in expenses:

            try:

                expense_date = (
                    datetime.strptime(
                        expense["date"],
                        "%Y-%m-%d"
                    )
                )


                if (
                    expense_date.year
                    == now.year
                    and
                    expense_date.month
                    == now.month
                ):

                    this_month += (
                        expense["amount"]
                    )


            except ValueError:

                continue


        # ----------------------------------------------------
        # MONTHLY TOTALS
        # ----------------------------------------------------

        monthly_totals = {}


        for expense in expenses:

            try:

                expense_date = (
                    datetime.strptime(
                        expense["date"],
                        "%Y-%m-%d"
                    )
                )


                month_key = (
                    expense_date.strftime(
                        "%Y-%m"
                    )
                )


                monthly_totals[
                    month_key
                ] = (

                    monthly_totals.get(
                        month_key,
                        0
                    )

                    +

                    expense["amount"]

                )


            except ValueError:

                continue


        # ----------------------------------------------------
        # MONTHLY AVERAGE
        # ----------------------------------------------------

        if monthly_totals:

            monthly_average = (
                sum(
                    monthly_totals.values()
                )
                /
                len(
                    monthly_totals
                )
            )

        else:

            monthly_average = 0


        # ----------------------------------------------------
        # HIGHEST EXPENSE
        # ----------------------------------------------------

        highest_expense = None


        if expenses:

            highest = max(
                expenses,
                key=lambda item:
                    item["amount"]
            )


            highest_expense = {

                "id":
                    highest["id"],

                "amount":
                    highest["amount"],

                "category":
                    highest["category"],

                "date":
                    display_date(
                        highest["date"]
                    )

            }


        # ----------------------------------------------------
        # RECENT EXPENSES
        # ----------------------------------------------------

        recent_expenses = sorted(
            expenses,
            key=lambda item:
                item["id"],
            reverse=True
        )[:5]


        # ----------------------------------------------------
        # RESPONSE
        # ----------------------------------------------------

        return jsonify({

            "total_expenses":
                total_expenses,

            "transaction_count":
                transaction_count,

            "average_expense":
                average_expense,

            "this_month":
                this_month,

            "monthly_average":
                monthly_average,

            "category_totals":
                category_totals,

            "year_totals":
                year_totals,

            "semester_totals":
                semester_totals,

            "payment_totals":
                payment_totals,

            "highest_expense":
                highest_expense,

            "recent_expenses":
                recent_expenses

        })


    except Exception as error:

        print(
            "STATS ERROR:",
            error
        )


        return jsonify({
            "error": str(error)
        }), 500


# ============================================================
# DOWNLOAD EXCEL
# ============================================================

@app.route(
    "/download-excel",
    methods=["GET"]
)
def download_excel():

    try:

        initialize_data()


        if not MASTER_EXCEL.exists():

            create_master_file()


        return send_file(

            MASTER_EXCEL,

            as_attachment=True,

            download_name=(
                "Student_Expense_Tracker.xlsx"
            ),

            mimetype=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            )

        )


    except Exception as error:

        print(
            "DOWNLOAD ERROR:",
            error
        )


        return jsonify({
            "error": str(error)
        }), 500


# ============================================================
# BACKUP CURRENT EXCEL
# ============================================================

@app.route(
    "/api/backup",
    methods=["POST"]
)
def create_backup():

    try:

        initialize_data()


        backup_file = (
            DATA_DIR /
            "Student_Expense_Tracker_Backup.xlsx"
        )


        with FILE_LOCK:

            source_workbook = load_workbook(
                MASTER_EXCEL
            )

            source_workbook.save(
                backup_file
            )

            source_workbook.close()


        return jsonify({

            "success":
                True,

            "message":
                "Backup created successfully.",

            "filename":
                backup_file.name

        })


    except Exception as error:

        print(
            "BACKUP ERROR:",
            error
        )


        return jsonify({
            "error": str(error)
        }), 500


# ============================================================
# HEALTH CHECK
# ============================================================

@app.route(
    "/api/health",
    methods=["GET"]
)
def health():

    return jsonify({

        "status":
            "ok",

        "database": "postgresql" if os.environ.get("DATABASE_URL") else "sqlite",
        "excel_export_available": True

    })


# ============================================================
# ERROR HANDLER
# ============================================================

@app.errorhandler(404)
def not_found(error):

    if request.path.startswith("/api/"):

        return jsonify({
            "error":
                "API endpoint not found."
        }), 404


    return error

# ============================================================
# SAVINGS
# ============================================================

SAVINGS_GOAL_COLUMNS = [
    "ID",
    "Goal Name",
    "Target Amount (₹)",
    "Target Date",
    "Description",
    "Created Date"
]

SAVINGS_HISTORY_COLUMNS = [
    "ID",
    "Goal ID",
    "Date",
    "Amount (₹)",
    "Note"
]


def clean_value(value):
    if value is None:
        return ""
    return str(value).strip()


def clean_number(value):
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def valid_date(value):
    return bool(normalize_date(value))


def initialize_savings_file():
    if SAVINGS_EXCEL.exists():
        return

    workbook = Workbook()
    goals_sheet = workbook.active
    goals_sheet.title = SAVINGS_GOALS_SHEET
    goals_sheet.append(SAVINGS_GOAL_COLUMNS)

    history_sheet = workbook.create_sheet(SAVINGS_HISTORY_SHEET)
    history_sheet.append(SAVINGS_HISTORY_COLUMNS)

    for worksheet in (goals_sheet, history_sheet):
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="7C3AED")
            cell.alignment = Alignment(horizontal="center")
        worksheet.freeze_panes = "A2"

    goals_sheet.column_dimensions["A"].width = 10
    goals_sheet.column_dimensions["B"].width = 28
    goals_sheet.column_dimensions["C"].width = 20
    goals_sheet.column_dimensions["D"].width = 16
    goals_sheet.column_dimensions["E"].width = 40
    goals_sheet.column_dimensions["F"].width = 16

    history_sheet.column_dimensions["A"].width = 10
    history_sheet.column_dimensions["B"].width = 12
    history_sheet.column_dimensions["C"].width = 16
    history_sheet.column_dimensions["D"].width = 18
    history_sheet.column_dimensions["E"].width = 40

    workbook.save(SAVINGS_EXCEL)


def read_savings_data():
    initialize_savings_file()

    with FILE_LOCK:
        workbook = load_workbook(SAVINGS_EXCEL, data_only=True)

        goals_sheet = workbook[SAVINGS_GOALS_SHEET]
        history_sheet = workbook[SAVINGS_HISTORY_SHEET]

        goals = []
        for row in goals_sheet.iter_rows(min_row=2, values_only=True):
            if not any(value is not None for value in row):
                continue
            values = list(row) + [None] * 6
            goals.append({
                "id": int(clean_number(values[0])),
                "name": clean_value(values[1]),
                "target_amount": clean_number(values[2]),
                "target_date": display_date(values[3]),
                "description": clean_value(values[4]),
                "created_date": display_date(values[5])
            })

        history = []
        for row in history_sheet.iter_rows(min_row=2, values_only=True):
            if not any(value is not None for value in row):
                continue
            values = list(row) + [None] * 5
            history.append({
                "id": int(clean_number(values[0])),
                "goal_id": int(clean_number(values[1])),
                "date": display_date(values[2]),
                "amount": clean_number(values[3]),
                "note": clean_value(values[4])
            })

        workbook.close()

    return goals, history


def save_savings_data(goals, history):
    with FILE_LOCK:
        workbook = Workbook()
        goals_sheet = workbook.active
        goals_sheet.title = SAVINGS_GOALS_SHEET
        goals_sheet.append(SAVINGS_GOAL_COLUMNS)

        for goal in goals:
            target_date = normalize_date(goal.get("target_date", ""))
            created_date = normalize_date(goal.get("created_date", ""))
            goals_sheet.append([
                int(goal["id"]),
                goal["name"],
                float(goal["target_amount"]),
                datetime.strptime(target_date, "%Y-%m-%d") if target_date else "",
                goal.get("description", ""),
                datetime.strptime(created_date, "%Y-%m-%d") if created_date else ""
            ])

        history_sheet = workbook.create_sheet(SAVINGS_HISTORY_SHEET)
        history_sheet.append(SAVINGS_HISTORY_COLUMNS)

        for item in history:
            date_value = normalize_date(item.get("date", ""))
            history_sheet.append([
                int(item["id"]),
                int(item["goal_id"]),
                datetime.strptime(date_value, "%Y-%m-%d") if date_value else "",
                float(item["amount"]),
                item.get("note", "")
            ])

        for worksheet in (goals_sheet, history_sheet):
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(fill_type="solid", fgColor="7C3AED")
                cell.alignment = Alignment(horizontal="center")
            worksheet.freeze_panes = "A2"

        for row in range(2, goals_sheet.max_row + 1):
            goals_sheet.cell(row, 3).number_format = '₹#,##0.00'
            goals_sheet.cell(row, 4).number_format = 'dd-mm-yyyy'
            goals_sheet.cell(row, 6).number_format = 'dd-mm-yyyy'

        for row in range(2, history_sheet.max_row + 1):
            history_sheet.cell(row, 3).number_format = 'dd-mm-yyyy'
            history_sheet.cell(row, 4).number_format = '₹#,##0.00'

        workbook.save(SAVINGS_EXCEL)


@app.route("/api/savings", methods=["GET"])
def api_get_savings():
    try:
        goals, history = read_savings_data()

        result_goals = []
        for goal in goals:
            saved_amount = sum(
                item["amount"] for item in history
                if item["goal_id"] == goal["id"]
            )
            target = float(goal["target_amount"])
            progress = (saved_amount / target * 100) if target > 0 else 0
            progress = min(max(progress, 0), 100)

            result_goals.append({
                **goal,
                "saved_amount": round(saved_amount, 2),
                "remaining_amount": round(max(target - saved_amount, 0), 2),
                "progress": round(progress, 1),
                "completed": saved_amount >= target
            })

        history_result = []
        for item in history:
            goal = next((g for g in goals if g["id"] == item["goal_id"]), None)
            history_result.append({
                **item,
                "goal_name": goal["name"] if goal else "Unknown Goal"
            })

        return jsonify({"goals": result_goals, "history": history_result})

    except Exception as error:
        print("GET SAVINGS ERROR:", error)
        return jsonify({"error": str(error)}), 500


@app.route("/api/savings/goals", methods=["POST"])
def api_create_savings_goal():
    data = request.get_json(silent=True) or {}
    name = clean_value(data.get("name"))
    target_amount = clean_number(data.get("target_amount"))
    target_date = clean_value(data.get("target_date"))
    description = clean_value(data.get("description"))

    if not name:
        return jsonify({"error": "Goal name is required."}), 400
    if target_amount <= 0:
        return jsonify({"error": "Target amount must be greater than zero."}), 400
    if target_date and not valid_date(target_date):
        return jsonify({"error": "Target date must be in DD-MM-YYYY format."}), 400

    goals, history = read_savings_data()
    new_id = max((goal["id"] for goal in goals), default=0) + 1

    goal = {
        "id": new_id,
        "name": name,
        "target_amount": target_amount,
        "target_date": target_date,
        "description": description,
        "created_date": datetime.now().strftime("%d-%m-%Y")
    }

    goals.append(goal)
    save_savings_data(goals, history)
    return jsonify({"success": True, "goal": goal}), 201


@app.route("/api/savings", methods=["POST"])
def api_add_savings():
    data = request.get_json(silent=True) or {}
    goal_id = int(clean_number(data.get("goal_id")))
    amount = clean_number(data.get("amount"))
    entered_date = clean_value(data.get("date"))
    note = clean_value(data.get("note"))

    if goal_id <= 0:
        return jsonify({"error": "Please select a savings goal."}), 400
    if amount <= 0:
        return jsonify({"error": "Savings amount must be greater than zero."}), 400
    if not valid_date(entered_date):
        return jsonify({"error": "Date must be in DD-MM-YYYY format."}), 400

    goals, history = read_savings_data()
    goal = next((g for g in goals if g["id"] == goal_id), None)
    if goal is None:
        return jsonify({"error": "Savings goal not found."}), 404

    saved_amount = sum(
        item["amount"] for item in history
        if item["goal_id"] == goal_id
    )
    remaining = max(goal["target_amount"] - saved_amount, 0)

    if amount > remaining:
        return jsonify({"error": "This amount is greater than the remaining goal amount."}), 400

    new_id = max((item["id"] for item in history), default=0) + 1
    item = {
        "id": new_id,
        "goal_id": goal_id,
        "date": entered_date,
        "amount": amount,
        "note": note
    }

    history.append(item)
    save_savings_data(goals, history)
    return jsonify({"success": True, "savings": item}), 201


@app.route("/api/savings/<int:savings_id>", methods=["DELETE"])
def api_delete_savings(savings_id):
    goals, history = read_savings_data()
    new_history = [item for item in history if item["id"] != savings_id]

    if len(new_history) == len(history):
        return jsonify({"error": "Savings record not found."}), 404

    save_savings_data(goals, new_history)
    return jsonify({"success": True})


@app.route("/api/savings/goals/<int:goal_id>", methods=["DELETE"])
def api_delete_savings_goal(goal_id):
    goals, history = read_savings_data()

    if not any(goal["id"] == goal_id for goal in goals):
        return jsonify({"error": "Savings goal not found."}), 404

    goals = [goal for goal in goals if goal["id"] != goal_id]
    history = [item for item in history if item["goal_id"] != goal_id]
    save_savings_data(goals, history)
    return jsonify({"success": True})



# ============================================================
# DATABASE STORAGE OVERRIDES
# ============================================================
# The desktop app historically used Excel as its live datastore. For a
# hosted Render deployment, live data is stored in SQLite locally or
# PostgreSQL when DATABASE_URL is configured. Excel remains available for
# downloads/backups.

def initialize_data():
    init_database()
    # Keep an Excel snapshot available for download/backup.
    # If the database was seeded with the user's old records, export those
    # records into the initial Excel snapshot as well.
    if not MASTER_EXCEL.exists():
        export_current_excel()


def get_expenses():
    initialize_data()
    return db_get_expenses()


def save_expenses(expenses):
    db_save_expenses(expenses)
    export_current_excel()


def read_savings_data():
    return db_read_savings_data()


def save_savings_data(goals, history):
    db_save_savings_data(goals, history)
    export_current_excel()


def export_current_excel():
    """Create an Excel snapshot from the database for download/backup."""
    initialize_data.__wrapped__ if hasattr(initialize_data, '__wrapped__') else None
    expenses = db_get_expenses()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    worksheet.append(HEADERS)
    for expense in expenses:
        date_text = normalize_date(expense.get("date", ""))
        excel_date = datetime.strptime(date_text, "%Y-%m-%d") if date_text else ""
        worksheet.append([
            int(expense.get("id", 0)), expense.get("academic_year", ""),
            expense.get("semester", ""), excel_date, expense.get("category", ""),
            float(expense.get("amount", 0)), expense.get("payment_method", ""),
            expense.get("description", "")
        ])
    format_excel(worksheet)
    workbook.save(MASTER_EXCEL)
    workbook.close()

# ============================================================
# START APPLICATION
# ============================================================

if __name__ == "__main__":

    initialize_data()


    print()
    print(
        "=================================================="
    )
    print(
        "        PERSONAL EXPENSE TRACKER"
    )
    print(
        "=================================================="
    )
    print(
        f"Project folder : {BASE_DIR}"
    )
    print(
        f"Data folder    : {DATA_DIR}"
    )
    print(
        f"Master Excel   : {MASTER_EXCEL}"
    )
    print(
        f"Old Excel      : {OLD_EXCEL}"
    )
    print(
        "Server         : http://127.0.0.1:5000"
    )
    print(
        "=================================================="
    )
    print()


    app.run(
        host="127.0.0.1",
        port=5000,
        debug=True
    )